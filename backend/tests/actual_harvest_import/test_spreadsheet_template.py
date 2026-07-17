from __future__ import annotations

import csv
import io
import zipfile

import openpyxl
import pytest

from backend.app.actual_harvest_import.spreadsheet_parser import (
    canonical_spreadsheet_headers,
    parse_csv,
    parse_xlsx,
)
from backend.app.actual_harvest_import.spreadsheet_template import (
    generate_csv_template,
    generate_xlsx_template,
)

pytestmark = [pytest.mark.unit, pytest.mark.contract]


def test_csv_and_xlsx_templates_have_the_same_derived_headers() -> None:
    csv_payload = generate_csv_template()
    reader = csv.reader(io.StringIO(csv_payload.decode("utf-8")))
    csv_headers = tuple(next(reader))

    workbook = openpyxl.load_workbook(io.BytesIO(generate_xlsx_template()), data_only=False)
    worksheet = workbook["actual_harvest"]
    xlsx_headers = tuple(cell.value for cell in worksheet[1])

    assert csv_headers == xlsx_headers == canonical_spreadsheet_headers()
    assert "import_received_at" not in csv_headers
    assert "ingested_at" not in csv_headers
    assert parse_csv(csv_payload).records == ()
    assert parse_xlsx(generate_xlsx_template()).records == ()


def test_templates_contain_no_formulas_macros_or_sample_values() -> None:
    xlsx_payload = generate_xlsx_template()
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_payload), data_only=False, keep_links=False)
    worksheet = workbook["actual_harvest"]
    assert workbook.sheetnames == ["actual_harvest"]
    assert all(cell.data_type != "f" for row in worksheet.iter_rows() for cell in row)
    assert all(
        cell.value is None or isinstance(cell.value, str)
        for row in worksheet.iter_rows()
        for cell in row
    )
    with zipfile.ZipFile(io.BytesIO(xlsx_payload)) as archive:
        assert not any(name.endswith("vbaProject.bin") for name in archive.namelist())


def test_template_logical_content_is_stable_and_round_trips() -> None:
    first = generate_xlsx_template()
    second = generate_xlsx_template()
    first_workbook = openpyxl.load_workbook(io.BytesIO(first), data_only=False)
    second_workbook = openpyxl.load_workbook(io.BytesIO(second), data_only=False)

    first_sheet = first_workbook["actual_harvest"]
    second_sheet = second_workbook["actual_harvest"]
    assert first_workbook.sheetnames == second_workbook.sheetnames
    assert tuple(cell.value for cell in first_sheet[1]) == tuple(
        cell.value for cell in second_sheet[1]
    )
    assert first_sheet.max_row == second_sheet.max_row == 1
    assert first_sheet.max_column == second_sheet.max_column
    assert parse_xlsx(first).diagnostics == parse_xlsx(second).diagnostics == ()
