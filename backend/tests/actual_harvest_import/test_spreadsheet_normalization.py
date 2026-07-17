from __future__ import annotations

import csv
import io
from datetime import date
from typing import Any

import openpyxl
import pytest

from backend.app.actual_harvest_import.enums import SpreadsheetDiagnosticCode
from backend.app.actual_harvest_import.schemas import ActualHarvestImportRecordInput
from backend.app.actual_harvest_import.spreadsheet_parser import (
    canonical_record_hash,
    canonical_record_payload,
    canonical_spreadsheet_headers,
    parse_csv,
    parse_xlsx,
)

pytestmark = [pytest.mark.unit, pytest.mark.contract]


def _row(**overrides: Any) -> dict[str, Any]:
    row: dict[str, Any] = {
        "external_logical_record_id": "logical-1",
        "external_revision_id": "revision-1",
        "source_system": "farm-picking-system",
        "external_batch_id": "batch-1",
        "harvest_business_date": "2026-03-01",
        "farm_code": "FARM-1",
        "subfarm_or_plot_code": "SUBFARM-1",
        "variety_code": "VARIETY-1",
        "actual_harvest_quantity_kg": "12.345678",
        "source_recorded_at": "",
        "source_recorded_at_authority_status": "MISSING",
        "source_recorded_at_authority_reference_or_null": "",
        "revision_number": "1",
        "record_status": "ACTIVE",
        "supersedes_external_revision_id": "",
        "season_code": "2026",
        "farm_timezone": "Asia/Shanghai",
        "revised_at": "",
        "finalized_at": "",
        "source_note": "note",
    }
    row.update(overrides)
    return row


def _csv(row: dict[str, Any]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=canonical_spreadsheet_headers(), lineterminator="\n")
    writer.writeheader()
    writer.writerow(row)
    return output.getvalue().encode("utf-8")


def _xlsx(row: dict[str, Any]) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "actual_harvest"
    for column, header in enumerate(canonical_spreadsheet_headers(), start=1):
        worksheet.cell(row=1, column=column, value=header)
        value = row[header]
        if header == "harvest_business_date":
            value = date.fromisoformat(value)
        elif value == "":
            value = None
        worksheet.cell(row=2, column=column, value=value)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_csv_and_xlsx_have_equivalent_canonical_payload_and_hash() -> None:
    csv_record = parse_csv(_csv(_row())).records[0]
    xlsx_record = parse_xlsx(_xlsx(_row())).records[0]

    assert canonical_record_payload(csv_record) == canonical_record_payload(xlsx_record)
    assert canonical_record_hash(csv_record) == canonical_record_hash(xlsx_record)


def test_canonical_headers_are_derived_from_existing_i1_schema() -> None:
    expected = (
        "external_logical_record_id",
        "external_revision_id",
        "source_system",
        "external_batch_id",
        "harvest_business_date",
        "farm_code",
        "subfarm_or_plot_code",
        "variety_code",
        "actual_harvest_quantity_kg",
        "source_recorded_at",
        "source_recorded_at_authority_status",
        "source_recorded_at_authority_reference_or_null",
        "revision_number",
        "record_status",
        "supersedes_external_revision_id",
        "season_code",
        "farm_timezone",
        "revised_at",
        "finalized_at",
        "source_note",
    )
    assert canonical_spreadsheet_headers() == expected
    assert "import_received_at" not in canonical_spreadsheet_headers()
    assert "ingested_at" not in canonical_spreadsheet_headers()
    assert ActualHarvestImportRecordInput.model_fields["source_row_number"].json_schema_extra == {
        "spreadsheet_importable": False
    }
    assert ActualHarvestImportRecordInput.model_fields["source_sheet_name"].json_schema_extra == {
        "spreadsheet_importable": False
    }


def test_row_and_sheet_provenance_is_diagnostic_not_business_hash_identity() -> None:
    csv_record = parse_csv(_csv(_row())).records[0]
    xlsx_record = parse_xlsx(_xlsx(_row())).records[0]

    assert csv_record.source_row_number == 2
    assert csv_record.source_sheet_name is None
    assert xlsx_record.source_row_number == 2
    assert xlsx_record.source_sheet_name == "actual_harvest"
    assert canonical_record_payload(csv_record) == canonical_record_payload(xlsx_record)


def test_empty_row_diagnostic_codes_are_typed_and_symmetric() -> None:
    csv_payload = _csv(_row()) + ("," * (len(canonical_spreadsheet_headers()) - 1) + "\n").encode()
    csv_result = parse_csv(csv_payload)
    assert csv_result.diagnostics[0].code == SpreadsheetDiagnosticCode.CSV_EMPTY_ROW_IGNORED

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "actual_harvest"
    for column, header in enumerate(canonical_spreadsheet_headers(), start=1):
        worksheet.cell(1, column, header)
    worksheet.cell(2, 1, "")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    xlsx_result = parse_xlsx(output.getvalue())
    assert xlsx_result.diagnostics[0].code == SpreadsheetDiagnosticCode.XLSX_EMPTY_ROW_IGNORED
