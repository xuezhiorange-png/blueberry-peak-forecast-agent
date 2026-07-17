from __future__ import annotations

import io
import re
import warnings
import zipfile
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import openpyxl
import pytest
from openpyxl.utils.datetime import MAC_EPOCH

from backend.app.actual_harvest_import.enums import ActualHarvestValidationErrorCode
from backend.app.actual_harvest_import.errors import ActualHarvestValidationError
from backend.app.actual_harvest_import.spreadsheet_parser import (
    canonical_spreadsheet_headers,
    parse_xlsx,
)
from backend.app.actual_harvest_import.spreadsheet_policy import SpreadsheetParserPolicy

pytestmark = [pytest.mark.unit, pytest.mark.contract]


def _row(**overrides: Any) -> dict[str, Any]:
    values: dict[str, Any] = {
        "external_logical_record_id": "logical-1",
        "external_revision_id": "revision-1",
        "source_system": "farm-picking-system",
        "external_batch_id": "batch-1",
        "harvest_business_date": date(2026, 3, 1),
        "farm_code": "FARM-1",
        "subfarm_or_plot_code": "SUBFARM-1",
        "variety_code": "VARIETY-1",
        "actual_harvest_quantity_kg": "12.345678",
        "source_recorded_at": None,
        "source_recorded_at_authority_status": "MISSING",
        "source_recorded_at_authority_reference_or_null": None,
        "revision_number": 1,
        "record_status": "ACTIVE",
        "supersedes_external_revision_id": None,
        "season_code": "2026",
        "farm_timezone": "Asia/Shanghai",
        "revised_at": None,
        "finalized_at": None,
        "source_note": "note",
    }
    values.update(overrides)
    return values


def _xlsx_bytes(
    rows: list[dict[str, Any]],
    *,
    sheet_name: str = "actual_harvest",
    headers: list[str] | None = None,
    extra_sheets: tuple[str, ...] = (),
    hidden_rows: tuple[int, ...] = (),
    merged_range: str | None = None,
    formula_cell: tuple[int, int] | None = None,
    hidden_columns: tuple[str, ...] = (),
) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    selected_headers = headers or list(canonical_spreadsheet_headers())
    for column, header in enumerate(selected_headers, start=1):
        worksheet.cell(row=1, column=column, value=header)
    for row_number, row in enumerate(rows, start=2):
        for column, header in enumerate(selected_headers, start=1):
            worksheet.cell(row=row_number, column=column, value=row.get(header))
    for row_number in hidden_rows:
        worksheet.row_dimensions[row_number].hidden = True
    for column in hidden_columns:
        worksheet.column_dimensions[column].hidden = True
    if merged_range is not None:
        worksheet.merge_cells(merged_range)
    if formula_cell is not None:
        worksheet.cell(row=formula_cell[0], column=formula_cell[1], value="=1+1")
    for name in extra_sheets:
        workbook.create_sheet(name)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _assert_code(payload: bytes, code: ActualHarvestValidationErrorCode) -> None:
    with pytest.raises(ActualHarvestValidationError) as captured:
        parse_xlsx(payload)
    assert captured.value.code == code
    assert captured.value.details is not None


def test_valid_xlsx_normalizes_date_decimal_and_sheet_provenance() -> None:
    result = parse_xlsx(_xlsx_bytes([_row()]))
    record = result.records[0]
    assert record.harvest_business_date == date(2026, 3, 1)
    assert record.actual_harvest_quantity_kg == Decimal("12.345678")
    assert record.source_row_number == 2
    assert record.source_sheet_name == "actual_harvest"


@pytest.mark.parametrize(
    ("payload_kwargs", "code"),
    [
        ({"sheet_name": "wrong"}, ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_MISSING),
        (
            {"extra_sheets": ("ACTUAL_HARVEST",)},
            ActualHarvestValidationErrorCode.XLSX_EXTRA_SHEET_FORBIDDEN,
        ),
        ({"formula_cell": (2, 9)}, ActualHarvestValidationErrorCode.XLSX_FORMULA_CELL_FORBIDDEN),
        ({"merged_range": "A2:B2"}, ActualHarvestValidationErrorCode.XLSX_MERGED_CELL_FORBIDDEN),
        ({"hidden_rows": (2,)}, ActualHarvestValidationErrorCode.XLSX_HIDDEN_ROW_FORBIDDEN),
    ],
)
def test_xlsx_structure_policy_is_fail_closed(
    payload_kwargs: dict[str, Any], code: ActualHarvestValidationErrorCode
) -> None:
    _assert_code(_xlsx_bytes([_row()], **payload_kwargs), code)


def test_xlsx_header_policy_rejects_unknown_and_duplicate_columns() -> None:
    headers = [*canonical_spreadsheet_headers(), "unknown_column"]
    _assert_code(
        _xlsx_bytes([_row()], headers=headers), ActualHarvestValidationErrorCode.UNKNOWN_FIELD
    )
    headers = [
        *canonical_spreadsheet_headers()[:2],
        canonical_spreadsheet_headers()[1],
        *canonical_spreadsheet_headers()[2:],
    ]
    _assert_code(
        _xlsx_bytes([_row()], headers=headers),
        ActualHarvestValidationErrorCode.CANONICAL_HEADER_DUPLICATE,
    )


def test_xlsx_case_insensitive_duplicate_canonical_sheet_is_rejected() -> None:
    payload = _xlsx_bytes([_row()], extra_sheets=("ACTUAL_HARVEST",))
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/workbook.xml":
                content = content.replace(b"ACTUAL_HARVEST1", b"ACTUAL_HARVEST")
            target.writestr(item, content)
    _assert_code(output.getvalue(), ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_DUPLICATE)


@pytest.mark.parametrize(
    "value",
    [1.0, datetime(2026, 3, 1, 12, 0), "03/01/2026"],
)
def test_ambiguous_or_binary_xlsx_values_are_rejected(value: Any) -> None:
    _assert_code(
        _xlsx_bytes([_row(harvest_business_date=value)]),
        ActualHarvestValidationErrorCode.INVALID_DATE,
    )


def test_xlsx_macro_content_is_rejected_before_workbook_load() -> None:
    payload = _xlsx_bytes([_row()])
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr("xl/vbaProject.bin", b"not executable")
    _assert_code(
        output.getvalue(), ActualHarvestValidationErrorCode.XLSX_EXECUTABLE_CONTENT_FORBIDDEN
    )


def test_xlsx_external_links_are_rejected_before_workbook_load() -> None:
    payload = _xlsx_bytes([_row()])
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr("xl/externalLinks/externalLink1.xml", b"<externalLink />")
    _assert_code(output.getvalue(), ActualHarvestValidationErrorCode.XLSX_EXTERNAL_LINK_FORBIDDEN)


def test_xlsx_serial_date_uses_the_versioned_date_cell_policy() -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "actual_harvest"
    headers = list(canonical_spreadsheet_headers())
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=header)
    values = _row()
    for column, header in enumerate(headers, start=1):
        value = values[header]
        if header == "harvest_business_date":
            value = 46082
        elif value == "":
            value = None
        worksheet.cell(row=2, column=column, value=value)
    worksheet.cell(
        row=2, column=headers.index("harvest_business_date") + 1
    ).number_format = "yyyy-mm-dd"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    result = parse_xlsx(output.getvalue())
    assert result.records[0].harvest_business_date == date(2026, 3, 1)


def test_xlsx_archive_limits_are_versioned_and_fail_closed() -> None:
    with pytest.raises(ActualHarvestValidationError) as captured:
        parse_xlsx(
            _xlsx_bytes([_row()]),
            policy=SpreadsheetParserPolicy(max_uncompressed_xlsx_size_bytes=1),
        )
    assert captured.value.code == ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED


def _rewrite_archive(payload: bytes, transform: Any) -> bytes:
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            target.writestr(item, transform(item.filename, source.read(item.filename)))
    return output.getvalue()


def _replace_quantity_lexical(payload: bytes, lexical: str) -> bytes:
    def transform(name: str, content: bytes) -> bytes:
        if name != "xl/worksheets/sheet1.xml":
            return content
        updated, count = re.subn(
            rb'<c[^>]*\br="I2"[^>]*>.*?</c>',
            b'<c r="I2" t="n"><v>' + lexical.encode() + b"</v></c>",
            content,
            count=1,
            flags=re.DOTALL,
        )
        assert count == 1
        return updated

    return _rewrite_archive(payload, transform)


def test_xlsx_external_defined_name_is_rejected() -> None:
    payload = _xlsx_bytes([_row()])
    workbook_xml = _rewrite_archive(
        payload,
        lambda name, content: (
            content.replace(
                b"</workbook>",
                b"<definedNames><definedName name=\"external\">='http://example.test/book.xlsx'!$A$1</definedName></definedNames></workbook>",
            )
            if name == "xl/workbook.xml"
            else content
        ),
    )
    _assert_code(workbook_xml, ActualHarvestValidationErrorCode.XLSX_EXTERNAL_LINK_FORBIDDEN)


@pytest.mark.parametrize(
    "relationship_type",
    [
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/activeXControl",
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package",
    ],
)
def test_xlsx_forbidden_relationships_are_rejected(relationship_type: str) -> None:
    payload = _xlsx_bytes([_row()])

    def transform(name: str, content: bytes) -> bytes:
        if name != "xl/_rels/workbook.xml.rels":
            return content
        return content.replace(
            b"</Relationships>",
            (
                f'<Relationship Type="{relationship_type}" Target="http://example.test/x" '
                'TargetMode="External" Id="bad" /></Relationships>'
            ).encode(),
        )

    _assert_code(
        _rewrite_archive(payload, transform),
        ActualHarvestValidationErrorCode.XLSX_EXTERNAL_LINK_FORBIDDEN,
    )


def test_xlsx_uppercase_relationship_part_is_rejected() -> None:
    payload = _xlsx_bytes([_row()])
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr(
            "xl/_rels/workbook.xml.RELS",
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
            b'externalLink" '
            b'Target="file:///tmp/outside.xlsx" TargetMode="External" Id="bad" />'
            b"</Relationships>",
        )
    _assert_code(output.getvalue(), ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID)


@pytest.mark.parametrize(
    "doctype",
    [
        b"<!DOCTYPE workbook>",
        b'<!DOCTYPE workbook [<!ENTITY x "safe">]>',
        b'<!DOCTYPE workbook SYSTEM "file:///tmp/secret">',
        b'<!DOCTYPE workbook PUBLIC "-//example//DTD//EN" "file:///tmp/secret">',
    ],
)
def test_xlsx_dtd_and_entity_declarations_are_rejected(doctype: bytes) -> None:
    payload = _xlsx_bytes([_row()])

    def transform(name: str, content: bytes) -> bytes:
        if name != "xl/workbook.xml":
            return content
        declaration_end = content.find(b"?>") + 2
        return content[:declaration_end] + doctype + content[declaration_end:]

    _assert_code(
        _rewrite_archive(payload, transform), ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID
    )


def test_xlsx_duplicate_parts_and_entry_count_are_rejected() -> None:
    payload = _xlsx_bytes([_row()])
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            target.writestr("xl/workbook.xml", source.read("xl/workbook.xml"))
    _assert_code(output.getvalue(), ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID)

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        entry_count = len(archive.infolist())
    assert parse_xlsx(
        payload, policy=SpreadsheetParserPolicy(max_xlsx_entry_count=entry_count)
    ).records
    with pytest.raises(ActualHarvestValidationError) as captured:
        parse_xlsx(payload, policy=SpreadsheetParserPolicy(max_xlsx_entry_count=entry_count - 1))
    assert captured.value.code == ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED


@pytest.mark.parametrize("unsafe_name", ["../escape", r"..\escape", "/absolute", "C:/drive"])
def test_xlsx_unsafe_part_names_are_rejected(unsafe_name: str) -> None:
    payload = _xlsx_bytes([_row()])
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload), "r") as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_STORED) as target,
    ):
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr(unsafe_name, b"x")
    _assert_code(output.getvalue(), ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID)


def test_xlsx_per_entry_compression_ratio_is_rejected() -> None:
    with pytest.raises(ActualHarvestValidationError) as captured:
        parse_xlsx(
            _xlsx_bytes([_row()]),
            policy=SpreadsheetParserPolicy(max_xlsx_compression_ratio=1),
        )
    assert captured.value.code == ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED


def test_xlsx_missing_or_malformed_required_parts_return_stable_errors() -> None:
    payload = _xlsx_bytes([_row()])
    required_parts = (
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    )
    for missing_part in required_parts:
        malformed = _rewrite_archive(
            payload,
            lambda name, content, missing_part=missing_part: (
                b"<broken" if name == missing_part else content
            ),
        )
        if missing_part == "[Content_Types].xml":
            output = io.BytesIO()
            with (
                zipfile.ZipFile(io.BytesIO(payload), "r") as source,
                zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
            ):
                for item in source.infolist():
                    if item.filename != missing_part:
                        target.writestr(item, source.read(item.filename))
            malformed = output.getvalue()
        _assert_code(malformed, ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID)


def test_xlsx_hidden_canonical_columns_are_rejected() -> None:
    _assert_code(
        _xlsx_bytes([_row()], hidden_columns=("I",)),
        ActualHarvestValidationErrorCode.XLSX_HIDDEN_COLUMN_FORBIDDEN,
    )


def test_xlsx_1904_epoch_serial_date_is_deterministic() -> None:
    workbook = openpyxl.Workbook()
    workbook.epoch = MAC_EPOCH
    worksheet = workbook.active
    worksheet.title = "actual_harvest"
    values = _row()
    headers = canonical_spreadsheet_headers()
    for column, header in enumerate(headers, start=1):
        value = values[header]
        if header == "harvest_business_date":
            value = 1
        elif value == "":
            value = None
        worksheet.cell(1, column, header)
        worksheet.cell(2, column, value)
    worksheet.cell(2, headers.index("harvest_business_date") + 1).number_format = "yyyy-mm-dd"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    result = parse_xlsx(output.getvalue())
    assert result.records[0].harvest_business_date == date(1904, 1, 2)


@pytest.mark.parametrize(
    ("lexical", "expected"),
    [
        ("0", Decimal("0")),
        ("12", Decimal("12")),
        ("12.5", Decimal("12.5")),
        ("12.345678", Decimal("12.345678")),
        ("0.000001", Decimal("0.000001")),
    ],
)
def test_xlsx_numeric_quantity_uses_raw_decimal_lexical(lexical: str, expected: Decimal) -> None:
    result = parse_xlsx(_replace_quantity_lexical(_xlsx_bytes([_row()]), lexical))
    assert result.records[0].actual_harvest_quantity_kg == expected


@pytest.mark.parametrize("lexical", ["12.3456789", "1000000000000", "-1", "1E2"])
def test_xlsx_numeric_quantity_contract_rejects_invalid_lexical(lexical: str) -> None:
    _assert_code(
        _replace_quantity_lexical(_xlsx_bytes([_row()]), lexical),
        ActualHarvestValidationErrorCode.INVALID_DECIMAL
        if not lexical.startswith("-")
        else ActualHarvestValidationErrorCode.NEGATIVE_QUANTITY,
    )


def test_csv_text_and_xlsx_numeric_quantity_are_equivalent() -> None:
    from backend.app.actual_harvest_import.spreadsheet_parser import parse_csv
    from backend.tests.actual_harvest_import.test_csv_parser import _csv_bytes

    csv_record = parse_csv(_csv_bytes([_row(actual_harvest_quantity_kg="12.5")])).records[0]
    xlsx_record = parse_xlsx(_replace_quantity_lexical(_xlsx_bytes([_row()]), "12.5")).records[0]
    assert csv_record.actual_harvest_quantity_kg == xlsx_record.actual_harvest_quantity_kg
