from __future__ import annotations

import csv
import hashlib
import io
import posixpath
import re
import xml.etree.ElementTree as ElementTree
import zipfile
from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO, NoReturn

import openpyxl
from openpyxl.utils.datetime import from_excel
from pydantic import BaseModel, ConfigDict, ValidationError

from backend.app.actual_harvest_import.enums import (
    ActualHarvestImportChannel,
    ActualHarvestValidationErrorCode,
    SpreadsheetDiagnosticCode,
)
from backend.app.actual_harvest_import.errors import ActualHarvestValidationError
from backend.app.actual_harvest_import.schemas import ActualHarvestImportRecordInput
from backend.app.actual_harvest_import.spreadsheet_policy import (
    DEFAULT_SPREADSHEET_POLICY,
    SpreadsheetParserPolicy,
)
from backend.app.rolling_backtest.canonical import canonical_json_dumps

_SERVER_GENERATED_FIELDS = frozenset({"import_received_at", "ingested_at"})
_DIAGNOSTIC_FIELDS = frozenset({"source_row_number", "source_sheet_name"})
_STRICT_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_STRICT_INTEGER_RE = re.compile(r"^[0-9]+$")
_STRICT_DECIMAL_RE = re.compile(r"^(?:0|[1-9]\d*)(?:\.\d{1,6})?$")
_XML_DECLARATION_RE = re.compile(rb"<!\s*(?:doctype|entity)\b", re.IGNORECASE)
_XML_EXTERNAL_TARGET_RE = re.compile(r"^(?:[a-z][a-z0-9+.-]*:|//|\\\\|[a-z]:[\\/])", re.IGNORECASE)
_XML_EXTERNAL_REFERENCE_RE = re.compile(
    r"(?:[a-z][a-z0-9+.-]*:|//|\\\\|[a-z]:[\\/])", re.IGNORECASE
)


class SpreadsheetParserError(ActualHarvestValidationError):
    """Typed parser error with safe source-location details."""


class SpreadsheetDiagnostic(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True, str_strip_whitespace=True)

    code: SpreadsheetDiagnosticCode
    source_channel: ActualHarvestImportChannel
    source_row_number: int | None = None
    source_sheet_name: str | None = None
    canonical_field_name: str | None = None
    detail: str


class SpreadsheetParseResult(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    source_channel: ActualHarvestImportChannel
    records: tuple[ActualHarvestImportRecordInput, ...]
    diagnostics: tuple[SpreadsheetDiagnostic, ...]


def canonical_spreadsheet_headers() -> tuple[str, ...]:
    """Return fields explicitly opted into the spreadsheet transport surface."""

    ordered_fields: list[tuple[int, str]] = []
    for field_name, field_info in ActualHarvestImportRecordInput.model_fields.items():
        metadata = field_info.json_schema_extra
        if not isinstance(metadata, dict) or metadata.get("spreadsheet_importable") is not True:
            continue
        order = metadata.get("spreadsheet_order")
        if not isinstance(order, int) or isinstance(order, bool):
            raise ValueError(f"spreadsheet field {field_name} has no stable order")
        ordered_fields.append((order, field_name))
    if len({order for order, _ in ordered_fields}) != len(ordered_fields):
        raise ValueError("spreadsheet field orders must be unique")
    return tuple(field_name for _, field_name in sorted(ordered_fields))


def canonical_record_payload(record: ActualHarvestImportRecordInput) -> dict[str, object]:
    """Return canonical business content without transport provenance."""

    return record.model_dump(mode="python", exclude=set(_DIAGNOSTIC_FIELDS))


def canonical_record_hash(record: ActualHarvestImportRecordInput) -> str:
    return hashlib.sha256(
        canonical_json_dumps(canonical_record_payload(record)).encode()
    ).hexdigest()


def parse_csv(
    source: bytes | bytearray | memoryview | BinaryIO | str | Path,
    *,
    policy: SpreadsheetParserPolicy = DEFAULT_SPREADSHEET_POLICY,
) -> SpreadsheetParseResult:
    raw = _read_source_bytes(source, policy=policy, channel=ActualHarvestImportChannel.CSV)
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        _fail(
            ActualHarvestValidationErrorCode.CSV_ENCODING_INVALID,
            ActualHarvestImportChannel.CSV,
            "CSV input is not valid UTF-8",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc

    previous_field_limit = csv.field_size_limit()
    csv.field_size_limit(policy.max_cell_text_length)
    try:
        reader = csv.reader(io.StringIO(text, newline=""), strict=True)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            _fail(
                ActualHarvestValidationErrorCode.CSV_STRUCTURE_INVALID,
                ActualHarvestImportChannel.CSV,
                "CSV must contain exactly one header row",
                details={"policy_version": policy.version},
            )
            raise AssertionError("unreachable") from exc
        headers = _validate_headers(raw_headers, ActualHarvestImportChannel.CSV, policy=policy)
        return _parse_rows(
            reader,
            headers,
            channel=ActualHarvestImportChannel.CSV,
            sheet_name=None,
            policy=policy,
        )
    except csv.Error as exc:
        code = (
            ActualHarvestValidationErrorCode.CSV_LIMIT_EXCEEDED
            if "field larger than field limit" in str(exc)
            else ActualHarvestValidationErrorCode.CSV_STRUCTURE_INVALID
        )
        _fail(
            code,
            ActualHarvestImportChannel.CSV,
            "CSV structure or cell-size policy is invalid",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc
    finally:
        csv.field_size_limit(previous_field_limit)


def parse_xlsx(
    source: bytes | bytearray | memoryview | BinaryIO | str | Path,
    *,
    policy: SpreadsheetParserPolicy = DEFAULT_SPREADSHEET_POLICY,
) -> SpreadsheetParseResult:
    raw = _read_source_bytes(source, policy=policy, channel=ActualHarvestImportChannel.XLSX)
    _validate_xlsx_archive(raw, policy)
    worksheet_part = _validate_xlsx_sheet_names(raw, policy)
    numeric_lexicals = _read_xlsx_numeric_lexicals(raw, worksheet_part, policy)
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw),
            data_only=False,
            read_only=False,
            keep_links=False,
        )
    except (OSError, ValueError, TypeError, IndexError, KeyError, zipfile.BadZipFile) as exc:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX workbook cannot be safely loaded",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc

    try:
        sheet_names = tuple(workbook.sheetnames)
        canonical_matches = tuple(
            name for name in sheet_names if name.casefold() == "actual_harvest"
        )
        if not canonical_matches:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_MISSING,
                ActualHarvestImportChannel.XLSX,
                "XLSX must contain the actual_harvest sheet",
                details={"policy_version": policy.version},
            )
        if len(canonical_matches) > 1:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_DUPLICATE,
                ActualHarvestImportChannel.XLSX,
                "XLSX contains duplicate canonical sheets",
                details={"policy_version": policy.version},
            )
        if len(sheet_names) > policy.max_sheet_count:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                ActualHarvestImportChannel.XLSX,
                "XLSX sheet count exceeds the parser policy",
                details={"policy_version": policy.version},
            )
        if canonical_matches[0] != "actual_harvest":
            _fail(
                ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_MISSING,
                ActualHarvestImportChannel.XLSX,
                "XLSX canonical sheet name must be exact",
                details={"policy_version": policy.version},
            )
        if len(sheet_names) != 1:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_EXTRA_SHEET_FORBIDDEN,
                ActualHarvestImportChannel.XLSX,
                "XLSX must contain only the canonical data sheet",
                details={"policy_version": policy.version},
            )

        worksheet = workbook["actual_harvest"]
        if (
            worksheet.max_row > policy.max_row_count
            or worksheet.max_column > policy.max_column_count
        ):
            _fail(
                ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                ActualHarvestImportChannel.XLSX,
                "XLSX data region exceeds the parser policy",
                sheet_name="actual_harvest",
                details={"policy_version": policy.version},
            )
        headers = _validate_headers(
            [cell.value for cell in worksheet[1]],
            ActualHarvestImportChannel.XLSX,
            sheet_name="actual_harvest",
            policy=policy,
        )
        _validate_xlsx_data_region(worksheet, headers, policy)

        def row_values() -> Iterable[list[object]]:
            for row_number in range(2, worksheet.max_row + 1):
                yield [
                    _xlsx_cell_value(
                        worksheet.cell(row=row_number, column=column),
                        field_name=header,
                        workbook=workbook,
                        row_number=row_number,
                        policy=policy,
                        numeric_lexicals=numeric_lexicals,
                    )
                    for column, header in enumerate(headers, start=1)
                ]

        return _parse_rows(
            row_values(),
            headers,
            channel=ActualHarvestImportChannel.XLSX,
            sheet_name="actual_harvest",
            policy=policy,
        )
    finally:
        workbook.close()


def _read_source_bytes(
    source: bytes | bytearray | memoryview | BinaryIO | str | Path,
    *,
    policy: SpreadsheetParserPolicy,
    channel: ActualHarvestImportChannel,
) -> bytes:
    limit_code = (
        ActualHarvestValidationErrorCode.CSV_LIMIT_EXCEEDED
        if channel == ActualHarvestImportChannel.CSV
        else ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED
    )
    invalid_code = (
        ActualHarvestValidationErrorCode.CSV_STRUCTURE_INVALID
        if channel == ActualHarvestImportChannel.CSV
        else ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID
    )
    try:
        if isinstance(source, bytes | bytearray | memoryview):
            if memoryview(source).nbytes > policy.max_file_size_bytes:
                _fail(
                    limit_code,
                    channel,
                    "spreadsheet file exceeds the parser policy",
                    details={"policy_version": policy.version},
                )
            return bytes(source)
        if isinstance(source, str | Path):
            with Path(source).open("rb") as stream:
                return _read_bounded_stream(stream, policy, channel, limit_code)
        return _read_bounded_stream(source, policy, channel, limit_code)
    except OSError as exc:
        _fail(
            invalid_code,
            channel,
            "spreadsheet source cannot be read",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc


def _read_bounded_stream(
    stream: BinaryIO,
    policy: SpreadsheetParserPolicy,
    channel: ActualHarvestImportChannel,
    limit_code: ActualHarvestValidationErrorCode,
) -> bytes:
    chunks: list[bytes] = []
    total = 0
    read_limit = policy.max_file_size_bytes + 1
    while total < read_limit:
        chunk = stream.read(min(64 * 1024, read_limit - total))
        if chunk in (b"", None):
            break
        if not isinstance(chunk, bytes):
            _fail(
                ActualHarvestValidationErrorCode.CSV_STRUCTURE_INVALID
                if channel == ActualHarvestImportChannel.CSV
                else ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
                channel,
                "spreadsheet input stream must produce bytes",
                details={"policy_version": policy.version},
            )
        chunks.append(chunk)
        total += len(chunk)
    if total > policy.max_file_size_bytes:
        _fail(
            limit_code,
            channel,
            "spreadsheet file exceeds the parser policy",
            details={"policy_version": policy.version},
        )
    return b"".join(chunks)


def _validate_headers(
    raw_headers: Sequence[object],
    channel: ActualHarvestImportChannel,
    *,
    policy: SpreadsheetParserPolicy,
    sheet_name: str | None = None,
) -> tuple[str, ...]:
    if not raw_headers or len(raw_headers) > policy.max_column_count:
        _fail(
            ActualHarvestValidationErrorCode.REQUIRED_FIELD_MISSING,
            channel,
            "spreadsheet must contain a canonical header row",
            sheet_name=sheet_name,
            details={"policy_version": policy.version},
        )
    headers: list[str] = []
    for value in raw_headers:
        if not isinstance(value, str):
            headers.append("")
        else:
            headers.append(value)
        if isinstance(value, str) and len(value) > policy.max_cell_text_length:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED
                if channel == ActualHarvestImportChannel.XLSX
                else ActualHarvestValidationErrorCode.CSV_LIMIT_EXCEEDED,
                channel,
                "header text exceeds the parser policy",
                sheet_name=sheet_name,
                details={"policy_version": policy.version},
            )
    if len(set(headers)) != len(headers):
        _fail(
            ActualHarvestValidationErrorCode.CANONICAL_HEADER_DUPLICATE,
            channel,
            "canonical headers must be unique",
            sheet_name=sheet_name,
            details={"policy_version": policy.version},
        )
    seen_normalized: dict[str, str] = {}
    for header in headers:
        normalized = header.strip().casefold()
        if normalized in seen_normalized and seen_normalized[normalized] != header:
            _fail(
                ActualHarvestValidationErrorCode.CANONICAL_HEADER_COLLISION,
                channel,
                "headers collide under diagnostic normalization; aliases are forbidden",
                sheet_name=sheet_name,
                canonical_field_name=header,
                details={"policy_version": policy.version},
            )
        seen_normalized[normalized] = header
    for header in headers:
        if header in _SERVER_GENERATED_FIELDS:
            _fail(
                ActualHarvestValidationErrorCode.SERVER_GENERATED_FIELD_SUPPLIED,
                channel,
                "server-generated fields must not be supplied by a spreadsheet",
                sheet_name=sheet_name,
                canonical_field_name=header,
                details={"policy_version": policy.version},
            )
    allowed = set(canonical_spreadsheet_headers())
    unknown = next((header for header in headers if header not in allowed), None)
    if unknown is not None:
        _fail(
            ActualHarvestValidationErrorCode.UNKNOWN_FIELD,
            channel,
            "spreadsheet contains an unknown canonical header",
            sheet_name=sheet_name,
            canonical_field_name=unknown or None,
            details={"policy_version": policy.version},
        )
    missing = next(
        (
            field_name
            for field_name in canonical_spreadsheet_headers()
            if ActualHarvestImportRecordInput.model_fields[field_name].is_required()
            and field_name not in headers
        ),
        None,
    )
    if missing is not None:
        _fail(
            ActualHarvestValidationErrorCode.REQUIRED_FIELD_MISSING,
            channel,
            "spreadsheet is missing a required canonical header",
            sheet_name=sheet_name,
            canonical_field_name=missing,
            details={"policy_version": policy.version},
        )
    return tuple(headers)


def _parse_rows(
    rows: Iterable[Sequence[object]],
    headers: tuple[str, ...],
    *,
    channel: ActualHarvestImportChannel,
    sheet_name: str | None,
    policy: SpreadsheetParserPolicy,
) -> SpreadsheetParseResult:
    records: list[ActualHarvestImportRecordInput] = []
    diagnostics: list[SpreadsheetDiagnostic] = []
    for offset, row in enumerate(rows, start=2):
        data_row_number = offset - 1
        if data_row_number > policy.max_row_count:
            _fail(
                ActualHarvestValidationErrorCode.CSV_LIMIT_EXCEEDED
                if channel == ActualHarvestImportChannel.CSV
                else ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                channel,
                "spreadsheet data row count exceeds the parser policy",
                row_number=offset,
                sheet_name=sheet_name,
                details={"policy_version": policy.version},
            )
        if len(row) != len(headers):
            _fail(
                ActualHarvestValidationErrorCode.CSV_STRUCTURE_INVALID
                if channel == ActualHarvestImportChannel.CSV
                else ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
                channel,
                "data row does not have the same number of cells as the header",
                row_number=offset,
                sheet_name=sheet_name,
                details={"policy_version": policy.version},
            )
        if all(_is_blank(value) for value in row):
            diagnostics.append(
                SpreadsheetDiagnostic(
                    code=(
                        SpreadsheetDiagnosticCode.CSV_EMPTY_ROW_IGNORED
                        if channel == ActualHarvestImportChannel.CSV
                        else SpreadsheetDiagnosticCode.XLSX_EMPTY_ROW_IGNORED
                    ),
                    source_channel=channel,
                    source_row_number=offset,
                    source_sheet_name=sheet_name,
                    detail="empty data row ignored by the canonical parser policy",
                )
            )
            continue
        payload = {header: value for header, value in zip(headers, row, strict=True)}
        records.append(
            _normalize_record(
                payload,
                channel=channel,
                row_number=offset,
                sheet_name=sheet_name,
                policy=policy,
            )
        )
    records.sort(key=_record_sort_key)
    return SpreadsheetParseResult(
        source_channel=channel,
        records=tuple(records),
        diagnostics=tuple(diagnostics),
    )


def _normalize_record(
    values: Mapping[str, object],
    *,
    channel: ActualHarvestImportChannel,
    row_number: int,
    sheet_name: str | None,
    policy: SpreadsheetParserPolicy,
) -> ActualHarvestImportRecordInput:
    payload: dict[str, object] = {}
    for field_name in canonical_spreadsheet_headers():
        value = values.get(field_name)
        if _is_blank(value):
            if ActualHarvestImportRecordInput.model_fields[field_name].is_required():
                _fail(
                    ActualHarvestValidationErrorCode.REQUIRED_FIELD_MISSING,
                    channel,
                    "required canonical field is missing",
                    row_number=row_number,
                    sheet_name=sheet_name,
                    canonical_field_name=field_name,
                    details={"policy_version": policy.version},
                )
            payload[field_name] = None
            continue
        if field_name == "actual_harvest_quantity_kg":
            payload[field_name] = _parse_decimal(
                value, channel, row_number, sheet_name, field_name, policy
            )
        elif field_name == "harvest_business_date":
            payload[field_name] = _parse_business_date(
                value, channel, row_number, sheet_name, field_name
            )
        elif field_name in {"source_recorded_at", "revised_at", "finalized_at"}:
            payload[field_name] = _parse_datetime(
                value, channel, row_number, sheet_name, field_name
            )
        elif field_name == "revision_number":
            payload[field_name] = _parse_integer(value, channel, row_number, sheet_name, field_name)
        elif isinstance(value, str):
            if "\x00" in value:
                _fail(
                    ActualHarvestValidationErrorCode.CSV_STRUCTURE_INVALID
                    if channel == ActualHarvestImportChannel.CSV
                    else ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
                    channel,
                    "cell text contains a NUL character",
                    row_number=row_number,
                    sheet_name=sheet_name,
                    canonical_field_name=field_name,
                    details={"policy_version": policy.version},
                )
            if len(value) > policy.max_cell_text_length:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED
                    if channel == ActualHarvestImportChannel.XLSX
                    else ActualHarvestValidationErrorCode.CSV_LIMIT_EXCEEDED,
                    channel,
                    "cell text exceeds the parser policy",
                    row_number=row_number,
                    sheet_name=sheet_name,
                    canonical_field_name=field_name,
                    details={"policy_version": policy.version},
                )
            payload[field_name] = value
        else:
            _fail(
                ActualHarvestValidationErrorCode.UNKNOWN_FIELD,
                channel,
                "cell type is not valid for the canonical field",
                row_number=row_number,
                sheet_name=sheet_name,
                canonical_field_name=field_name,
                details={"policy_version": policy.version},
            )
    payload["source_row_number"] = row_number
    payload["source_sheet_name"] = sheet_name
    try:
        return ActualHarvestImportRecordInput.model_validate(payload)
    except ValidationError as exc:
        first_error = exc.errors()[0]
        location = first_error.get("loc", ())
        field_name = str(location[0]) if location else "unknown"
        code = _validation_error_code(field_name, str(first_error.get("type", "invalid")))
        _fail(
            code,
            channel,
            "canonical record failed the existing I1 schema validation",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={
                "policy_version": policy.version,
                "validation_type": first_error.get("type", "invalid"),
            },
        )
        raise AssertionError("unreachable") from exc


def _parse_decimal(
    value: object,
    channel: ActualHarvestImportChannel,
    row_number: int,
    sheet_name: str | None,
    field_name: str,
    policy: SpreadsheetParserPolicy,
) -> Decimal:
    if isinstance(value, bool) or isinstance(value, float):
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "quantity must not use native float or bool",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if isinstance(value, Decimal | int):
        text = str(value)
    elif isinstance(value, str):
        text = value.strip()
    else:
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "quantity must be a decimal string or exact integer",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if text.startswith("-"):
        _fail(
            ActualHarvestValidationErrorCode.NEGATIVE_QUANTITY,
            channel,
            "actual harvest quantity must be non-negative",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if not _STRICT_DECIMAL_RE.fullmatch(text):
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "quantity must use a finite non-scientific decimal within NUMERIC(18,6)",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    try:
        decimal_value = Decimal(text)
    except InvalidOperation as exc:
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "quantity is not a valid Decimal",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc
    if not decimal_value.is_finite():
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "quantity must be finite",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if decimal_value < 0:
        _fail(
            ActualHarvestValidationErrorCode.NEGATIVE_QUANTITY,
            channel,
            "actual harvest quantity must be non-negative",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if len(text.split(".", maxsplit=1)[0]) > 12:
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "quantity exceeds NUMERIC(18,6) integer precision",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    return decimal_value


def _parse_integer(
    value: object,
    channel: ActualHarvestImportChannel,
    row_number: int,
    sheet_name: str | None,
    field_name: str,
) -> int:
    if isinstance(value, bool):
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DECIMAL,
            channel,
            "integer field must not be bool",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={},
        )
    if isinstance(value, int):
        return value
    if isinstance(value, str) and _STRICT_INTEGER_RE.fullmatch(value.strip()):
        return int(value)
    _fail(
        ActualHarvestValidationErrorCode.INVALID_DECIMAL,
        channel,
        "revision_number must be a strict integer",
        row_number=row_number,
        sheet_name=sheet_name,
        canonical_field_name=field_name,
        details={},
    )
    raise AssertionError("unreachable")


def _parse_business_date(
    value: object,
    channel: ActualHarvestImportChannel,
    row_number: int,
    sheet_name: str | None,
    field_name: str,
) -> date:
    if isinstance(value, datetime):
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DATE,
            channel,
            "harvest_business_date must be a farm-local date, not a datetime",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={},
        )
    if isinstance(value, date):
        return value
    if isinstance(value, str) and _STRICT_DATE_RE.fullmatch(value.strip()):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            pass
    _fail(
        ActualHarvestValidationErrorCode.INVALID_DATE,
        channel,
        "harvest_business_date must use YYYY-MM-DD without locale inference",
        row_number=row_number,
        sheet_name=sheet_name,
        canonical_field_name=field_name,
        details={},
    )
    raise AssertionError("unreachable")


def _parse_datetime(
    value: object,
    channel: ActualHarvestImportChannel,
    row_number: int,
    sheet_name: str | None,
    field_name: str,
) -> datetime:
    if not isinstance(value, str):
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DATETIME,
            channel,
            "source datetime must be an explicit timezone-aware ISO-8601 string",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={},
        )
    text = value.strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError as exc:
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DATETIME,
            channel,
            "source datetime is not valid ISO-8601",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={},
        )
        raise AssertionError("unreachable") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        _fail(
            ActualHarvestValidationErrorCode.INVALID_DATETIME,
            channel,
            "source datetime must include an explicit timezone",
            row_number=row_number,
            sheet_name=sheet_name,
            canonical_field_name=field_name,
            details={},
        )
    return parsed


def _xlsx_cell_value(
    cell: Any,
    *,
    field_name: str,
    workbook: Any,
    row_number: int,
    policy: SpreadsheetParserPolicy,
    numeric_lexicals: Mapping[str, str],
) -> object:
    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
        _fail(
            ActualHarvestValidationErrorCode.XLSX_FORMULA_CELL_FORBIDDEN,
            ActualHarvestImportChannel.XLSX,
            "formula cells are forbidden in the canonical data region",
            row_number=row_number,
            sheet_name="actual_harvest",
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if isinstance(cell.value, str) and len(cell.value) > policy.max_cell_text_length:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
            ActualHarvestImportChannel.XLSX,
            "cell text exceeds the parser policy",
            row_number=row_number,
            sheet_name="actual_harvest",
            canonical_field_name=field_name,
            details={"policy_version": policy.version},
        )
    if (
        field_name == "actual_harvest_quantity_kg"
        and cell.data_type == "n"
        and isinstance(cell.value, int | float)
    ):
        lexical = numeric_lexicals.get(cell.coordinate)
        if lexical is None:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
                ActualHarvestImportChannel.XLSX,
                "numeric cell has no canonical OOXML value",
                row_number=row_number,
                sheet_name="actual_harvest",
                canonical_field_name=field_name,
                details={"policy_version": policy.version},
            )
        return lexical
    if (
        field_name == "harvest_business_date"
        and cell.is_date
        and isinstance(cell.value, int | float)
    ):
        converted = from_excel(cell.value, workbook.epoch)
        if isinstance(converted, datetime):
            if converted.time() != time.min:
                _fail(
                    ActualHarvestValidationErrorCode.INVALID_DATE,
                    ActualHarvestImportChannel.XLSX,
                    "spreadsheet serial datetime has a non-midnight time",
                    row_number=row_number,
                    sheet_name="actual_harvest",
                    canonical_field_name=field_name,
                    details={"policy_version": policy.version},
                )
            return converted.date()
        return converted
    if field_name == "harvest_business_date" and isinstance(cell.value, datetime):
        format_text = cell.number_format.lower()
        if any(marker in format_text for marker in ("h", "s")):
            _fail(
                ActualHarvestValidationErrorCode.INVALID_DATE,
                ActualHarvestImportChannel.XLSX,
                "datetime cells are not accepted as harvest business dates",
                row_number=row_number,
                sheet_name="actual_harvest",
                canonical_field_name=field_name,
                details={"policy_version": policy.version},
            )
        if cell.value.time() == time.min:
            return cell.value.date()
    return cell.value


def _validate_xlsx_data_region(
    worksheet: Any,
    headers: tuple[str, ...],
    policy: SpreadsheetParserPolicy,
) -> None:
    for row_number in range(1, worksheet.max_row + 1):
        if worksheet.row_dimensions[row_number].hidden:
            _fail(
                ActualHarvestValidationErrorCode.XLSX_HIDDEN_ROW_FORBIDDEN,
                ActualHarvestImportChannel.XLSX,
                "hidden rows are forbidden in the canonical data region",
                row_number=row_number,
                sheet_name="actual_harvest",
                details={"policy_version": policy.version},
            )
    for _column_letter, dimension in worksheet.column_dimensions.items():
        if dimension.hidden and (dimension.min or 1) <= len(headers):
            _fail(
                ActualHarvestValidationErrorCode.XLSX_HIDDEN_COLUMN_FORBIDDEN,
                ActualHarvestImportChannel.XLSX,
                "hidden columns are forbidden in the canonical data region",
                sheet_name="actual_harvest",
                details={"policy_version": policy.version},
            )
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= worksheet.max_row and merged_range.min_col <= len(headers):
            _fail(
                ActualHarvestValidationErrorCode.XLSX_MERGED_CELL_FORBIDDEN,
                ActualHarvestImportChannel.XLSX,
                "merged cells are forbidden in the canonical data region",
                row_number=merged_range.min_row,
                sheet_name="actual_harvest",
                details={"policy_version": policy.version},
            )


def _validate_xlsx_archive(raw: bytes, policy: SpreadsheetParserPolicy) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            infos = archive.infolist()
            if len(infos) > policy.max_xlsx_entry_count:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX entry count exceeds the parser policy",
                    details={"policy_version": policy.version},
                )
            seen_names: set[str] = set()
            seen_casefolded_names: set[str] = set()
            total_uncompressed = 0
            total_compressed = 0
            for item in infos:
                normalized_name = _normalize_zip_entry_name(item.filename, policy)
                casefolded_name = normalized_name.casefold()
                if normalized_name in seen_names or casefolded_name in seen_casefolded_names:
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX archive contains duplicate part names",
                        details={"policy_version": policy.version},
                    )
                seen_names.add(normalized_name)
                seen_casefolded_names.add(casefolded_name)
                lowered_name = normalized_name.lower()
                if lowered_name.startswith("xl/externallinks/") or "externallink" in lowered_name:
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_EXTERNAL_LINK_FORBIDDEN,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX external links are forbidden",
                        details={"policy_version": policy.version},
                    )
                if (
                    lowered_name.endswith("vbaproject.bin")
                    or lowered_name.startswith("xl/activex/")
                    or lowered_name.startswith("xl/embeddings/")
                    or lowered_name.startswith("xl/macrosheets/")
                    or lowered_name.startswith("customui/")
                ):
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_EXECUTABLE_CONTENT_FORBIDDEN,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX executable content is forbidden",
                        details={"policy_version": policy.version},
                    )
                if lowered_name.endswith(".rels"):
                    relationship_root = _safe_xml_from_bytes(archive.read(item.filename), policy)
                    _reject_forbidden_relationships(relationship_root, policy)
                elif lowered_name.endswith(".xml"):
                    _safe_xml_from_bytes(archive.read(item.filename), policy)
                if item.file_size > 0 and item.compress_size == 0:
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX entry has invalid compression metadata",
                        details={"policy_version": policy.version},
                    )
                if (
                    item.file_size > 0
                    and item.file_size > item.compress_size * policy.max_xlsx_compression_ratio
                ):
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX entry compression ratio exceeds the parser policy",
                        details={"policy_version": policy.version},
                    )
                total_uncompressed += item.file_size
                total_compressed += item.compress_size
            content_types_item = next(
                (item for item in infos if item.filename == "[Content_Types].xml"), None
            )
            if content_types_item is None:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX content types part is missing",
                    details={"policy_version": policy.version},
                )
            content_types = _safe_xml_from_bytes(archive.read(content_types_item.filename), policy)
            for element in content_types.iter():
                values = " ".join(str(value).lower() for value in element.attrib.values())
                if any(
                    token in values
                    for token in ("macroenabled", "vbaproject", "activex", "oleobject")
                ):
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_EXECUTABLE_CONTENT_FORBIDDEN,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX executable content types are forbidden",
                        details={"policy_version": policy.version},
                    )
            if total_uncompressed > policy.max_uncompressed_xlsx_size_bytes:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX uncompressed size exceeds the parser policy",
                    details={"policy_version": policy.version},
                )
            if total_uncompressed > 0 and total_compressed == 0:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX aggregate compression metadata is invalid",
                    details={"policy_version": policy.version},
                )
            if total_uncompressed > total_compressed * policy.max_xlsx_compression_ratio:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_LIMIT_EXCEEDED,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX aggregate compression ratio exceeds the parser policy",
                    details={"policy_version": policy.version},
                )
    except ActualHarvestValidationError:
        raise
    except (OSError, KeyError, IndexError, ValueError, zipfile.BadZipFile) as exc:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX archive is invalid",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc


def _normalize_zip_entry_name(name: str, policy: SpreadsheetParserPolicy) -> str:
    if not name or "\x00" in name or "\\" in name:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX archive contains an unsafe part name",
            details={"policy_version": policy.version},
        )
    if name.startswith("/") or re.match(r"^[A-Za-z]:", name):
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX archive contains an absolute part name",
            details={"policy_version": policy.version},
        )
    raw_parts = name.split("/")
    if any(part in {"", ".", ".."} for part in raw_parts):
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX archive contains an ambiguous part name",
            details={"policy_version": policy.version},
        )
    normalized = posixpath.normpath(name)
    if normalized in {"", "."} or normalized.startswith("../") or normalized == "..":
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX archive contains a traversal part name",
            details={"policy_version": policy.version},
        )
    return normalized


def _safe_xml_from_bytes(raw_xml: bytes, policy: SpreadsheetParserPolicy) -> ElementTree.Element:
    if _XML_DECLARATION_RE.search(raw_xml):
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX XML DTD and entity declarations are forbidden",
            details={"policy_version": policy.version},
        )
    try:
        return ElementTree.fromstring(raw_xml)
    except (ElementTree.ParseError, ValueError, TypeError) as exc:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX XML part is invalid",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", maxsplit=1)[-1]


def _reject_forbidden_relationships(
    root: ElementTree.Element, policy: SpreadsheetParserPolicy
) -> None:
    for element in root.iter():
        if _xml_local_name(element.tag) != "Relationship":
            continue
        attributes = {
            key.rsplit("}", maxsplit=1)[-1].lower(): value for key, value in element.attrib.items()
        }
        relationship_type = attributes.get("type", "").lower()
        relationship_kind = relationship_type.rsplit("/", maxsplit=1)[-1]
        target = attributes.get("target", "")
        target_mode = attributes.get("targetmode", "").lower()
        if (
            target_mode == "external"
            or _XML_EXTERNAL_TARGET_RE.match(target) is not None
            or any(
                token in relationship_kind
                for token in ("externallink", "activex", "oleobject", "vba", "package", "dde")
            )
            or "|" in target
        ):
            _fail(
                ActualHarvestValidationErrorCode.XLSX_EXTERNAL_LINK_FORBIDDEN
                if target_mode == "external"
                or "external" in relationship_type
                or _XML_EXTERNAL_TARGET_RE.match(target) is not None
                else ActualHarvestValidationErrorCode.XLSX_EXECUTABLE_CONTENT_FORBIDDEN,
                ActualHarvestImportChannel.XLSX,
                "XLSX external or executable relationships are forbidden",
                details={"policy_version": policy.version},
            )


def _validate_xlsx_sheet_names(raw: bytes, policy: SpreadsheetParserPolicy) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            workbook_xml = _safe_xml_from_bytes(archive.read("xl/workbook.xml"), policy)
            relationships_xml = _safe_xml_from_bytes(
                archive.read("xl/_rels/workbook.xml.rels"), policy
            )
            _reject_forbidden_relationships(relationships_xml, policy)
            relationship_targets = {
                element.attrib.get("Id", ""): element.attrib.get("Target", "")
                for element in relationships_xml
                if _xml_local_name(element.tag) == "Relationship"
            }
            sheet_elements = [
                element
                for element in workbook_xml.iter()
                if _xml_local_name(element.tag) == "sheet"
            ]
            sheet_names = tuple(element.attrib.get("name", "") for element in sheet_elements)
            canonical_matches = tuple(
                name for name in sheet_names if name.casefold() == "actual_harvest"
            )
            if len(canonical_matches) > 1:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_DUPLICATE,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX contains duplicate canonical sheets",
                    details={"policy_version": policy.version},
                )
            if not canonical_matches:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_MISSING,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX must contain the actual_harvest sheet",
                    details={"policy_version": policy.version},
                )
            if canonical_matches[0] != "actual_harvest":
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_CANONICAL_SHEET_MISSING,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX canonical sheet name must be exact",
                    details={"policy_version": policy.version},
                )
            if len(sheet_names) > policy.max_sheet_count:
                _fail(
                    ActualHarvestValidationErrorCode.XLSX_EXTRA_SHEET_FORBIDDEN,
                    ActualHarvestImportChannel.XLSX,
                    "XLSX must contain only the canonical data sheet",
                    details={"policy_version": policy.version},
                )
            canonical_sheet = next(
                element
                for element in sheet_elements
                if element.attrib.get("name") == "actual_harvest"
            )
            relationship_id = next(
                value
                for key, value in canonical_sheet.attrib.items()
                if key.rsplit("}", maxsplit=1)[-1] == "id"
            )
            target = relationship_targets.get(relationship_id)
            if not target:
                raise KeyError("canonical worksheet relationship is missing")
            worksheet_part = (
                posixpath.normpath(target.lstrip("/"))
                if target.startswith("/")
                else posixpath.normpath(posixpath.join("xl", target))
            )
            if worksheet_part not in archive.namelist():
                raise KeyError("canonical worksheet part is missing")
            for element in workbook_xml.iter():
                if _xml_local_name(element.tag) != "definedName":
                    continue
                value = (element.text or "").lower()
                if (
                    "[" in value
                    or "]" in value
                    or "|" in value
                    or _XML_EXTERNAL_REFERENCE_RE.search(value) is not None
                ):
                    _fail(
                        ActualHarvestValidationErrorCode.XLSX_EXTERNAL_LINK_FORBIDDEN,
                        ActualHarvestImportChannel.XLSX,
                        "XLSX external defined names are forbidden",
                        details={"policy_version": policy.version},
                    )
            return worksheet_part
    except ActualHarvestValidationError:
        raise
    except (OSError, KeyError, IndexError, StopIteration, ValueError, zipfile.BadZipFile) as exc:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX workbook metadata is invalid",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc


def _read_xlsx_numeric_lexicals(
    raw: bytes, worksheet_part: str, policy: SpreadsheetParserPolicy
) -> dict[str, str]:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            worksheet_xml = _safe_xml_from_bytes(archive.read(worksheet_part), policy)
    except ActualHarvestValidationError:
        raise
    except (OSError, KeyError, IndexError, ValueError, zipfile.BadZipFile) as exc:
        _fail(
            ActualHarvestValidationErrorCode.XLSX_ARCHIVE_INVALID,
            ActualHarvestImportChannel.XLSX,
            "XLSX canonical worksheet part is invalid",
            details={"policy_version": policy.version},
        )
        raise AssertionError("unreachable") from exc
    values: dict[str, str] = {}
    for element in worksheet_xml.iter():
        if _xml_local_name(element.tag) != "c":
            continue
        cell_type = element.attrib.get("t", "n")
        if cell_type not in {"n", ""}:
            continue
        reference = element.attrib.get("r")
        value_element = next(
            (child for child in element if _xml_local_name(child.tag) == "v"), None
        )
        if reference and value_element is not None:
            values[reference] = value_element.text or ""
    return values


def _validation_error_code(field_name: str, error_type: str) -> ActualHarvestValidationErrorCode:
    if field_name == "harvest_business_date":
        return ActualHarvestValidationErrorCode.INVALID_DATE
    if field_name in {"source_recorded_at", "revised_at", "finalized_at"}:
        return ActualHarvestValidationErrorCode.INVALID_DATETIME
    if field_name == "farm_timezone":
        return ActualHarvestValidationErrorCode.INVALID_TIMEZONE
    if field_name == "actual_harvest_quantity_kg":
        return ActualHarvestValidationErrorCode.INVALID_DECIMAL
    if field_name == "record_status":
        return ActualHarvestValidationErrorCode.INVALID_RECORD_STATUS
    if error_type == "missing":
        return ActualHarvestValidationErrorCode.REQUIRED_FIELD_MISSING
    return ActualHarvestValidationErrorCode.UNKNOWN_FIELD


def _record_sort_key(record: ActualHarvestImportRecordInput) -> tuple[object, ...]:
    return (
        record.source_system,
        record.external_batch_id,
        record.external_logical_record_id,
        record.external_revision_id,
        record.revision_number,
        record.harvest_business_date,
        record.farm_code,
        record.subfarm_or_plot_code,
        record.variety_code,
    )


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _fail(
    code: ActualHarvestValidationErrorCode,
    channel: ActualHarvestImportChannel,
    message: str,
    *,
    row_number: int | None = None,
    sheet_name: str | None = None,
    canonical_field_name: str | None = None,
    details: dict[str, object],
) -> NoReturn:
    safe_details = {
        "source_channel": channel.value,
        "source_row_number": row_number,
        "source_sheet_name": sheet_name,
        "canonical_field_name": canonical_field_name,
        **details,
    }
    raise SpreadsheetParserError(
        code,
        message,
        field_path=canonical_field_name,
        details=safe_details,
    )
